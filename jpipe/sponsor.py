"""Company-level sponsorship records.

Two data sources, both public and free:
  1. DOL OFLC LCA disclosure files (quarterly xlsx/csv) -- H-1B filing level, largest volume
     https://www.dol.gov/agencies/eta/foreign-labor/performance
  2. USCIS H-1B Employer Data Hub (csv export) -- approval level, stronger evidence
     that petitions actually got approved
     https://www.uscis.gov/tools/reports-and-studies/h-1b-employer-data-hub

Usage:
    python jobpipe.py sponsor-ingest ~/Downloads/LCA_Disclosure_FY2025_Q3.xlsx

When no data has been ingested, a company's sponsor flag falls back to the seed
guess in companies.yaml.
"""

from __future__ import annotations

import csv
from pathlib import Path

from .db import connect, normalize_company, now

# Calibrated against the actual USCIS FY2022+FY2023 distribution (62599 companies):
#   101+  only 163 companies (top 0.3%, almost all outsourcing giants)
#   26+   ~820 companies <- the actual level of a "reliably sponsoring" tech company
#         (Stripe 71, Databricks 29)
#   3+    ~8800 companies
# So setting heavy at 100 would be wrong: it would drop every normal tech company
# down to yes.
HEAVY_THRESHOLD = 25
YES_THRESHOLD = 3  # below 3 counts as low: has sponsored, but not as a routine practice

# Possible column names per data source (case-insensitive)
EMPLOYER_COLS = ["employer_name", "employer", "employer legal business name", "lca_case_employer_name"]
COUNT_COLS = ["initial approval", "initial approvals", "continuing approval"]


def _find_col(header: list[str], candidates: list[str]) -> int | None:
    lowered = [(h or "").strip().lower() for h in header]
    for cand in candidates:
        if cand in lowered:
            return lowered.index(cand)
    return None


def _rows_from_csv(path: Path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        yield from csv.reader(f)


def _rows_from_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Reading xlsx requires openpyxl: pip install openpyxl\n"
            "(or save it as csv with Excel/Numbers and ingest that instead)"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
        yield ["" if c is None else str(c) for c in row]


def ingest(path_str: str, source: str | None = None) -> tuple[int, int]:
    """Ingest one disclosure file; returns (rows read, companies written)."""
    path = Path(path_str).expanduser()
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    rows = _rows_from_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") else _rows_from_csv(path)
    rows = iter(rows)
    try:
        header = next(rows)
    except StopIteration:
        raise SystemExit("File is empty")

    emp_idx = _find_col(header, EMPLOYER_COLS)
    if emp_idx is None:
        raise SystemExit(
            f"Employer name column not found. The file's header is:\n  {', '.join(header[:25])}\n"
            f"Supported column names: {EMPLOYER_COLS}"
        )
    cnt_idx = _find_col(header, COUNT_COLS)  # USCIS hub has an approvals column; LCA doesn't, so count rows

    tally: dict[str, tuple[str, int]] = {}
    n_rows = 0
    for row in rows:
        if emp_idx >= len(row):
            continue
        raw = (row[emp_idx] or "").strip()
        if not raw:
            continue
        n_rows += 1
        key = normalize_company(raw)
        if not key:
            continue
        inc = 1
        if cnt_idx is not None and cnt_idx < len(row):
            try:
                inc = int(float(row[cnt_idx] or 0))
            except ValueError:
                inc = 1
        display, cur = tally.get(key, (raw, 0))
        tally[key] = (display, cur + inc)

    src = source or path.name
    ts = now()
    conn = connect()
    with conn:
        for key, (display, count) in tally.items():
            conn.execute(
                """INSERT INTO sponsor_records(company_norm, display_name, count, source, updated_at)
                   VALUES(?,?,?,?,?)
                   ON CONFLICT(company_norm) DO UPDATE SET
                       count = sponsor_records.count + excluded.count,
                       source = excluded.source,
                       updated_at = excluded.updated_at""",
                (key, display, count, src, ts),
            )
    conn.close()
    return n_rows, len(tally)


def has_data(conn) -> bool:
    return conn.execute("SELECT 1 FROM sponsor_records LIMIT 1").fetchone() is not None


def _count_for(conn, name: str) -> int | None:
    key = normalize_company(name)
    if not key:
        return None
    row = conn.execute("SELECT count FROM sponsor_records WHERE company_norm=?", (key,)).fetchone()
    if row:
        return row["count"]
    if len(key) >= 5:
        # Prefix match when the names don't line up exactly ("stripe" vs "stripe payments inc")
        row = conn.execute(
            "SELECT SUM(count) AS count FROM sponsor_records WHERE company_norm LIKE ?",
            (key + " %",),
        ).fetchone()
        if row and row["count"] is not None:
            return row["count"]
    return None


def lookup(conn, company: str, fallback: str = "unknown", aliases: list[str] | None = None) -> str:
    """Return heavy | yes | low | none | unknown. If no data has ever been ingested,
    fall back to `fallback` (the seed guess from the yaml).

    aliases are legal entity names -- a brand name and its legal name very often
    don't match (Instacart's legal entity is MAPLEBEAR INC, Airtable is FORMAGRID
    INC, SoFi is SOCIAL FINANCE INC); without aliases these would be misclassified
    as "no records found".
    """
    if not has_data(conn):
        return fallback

    best = None
    for name in [company, *(aliases or [])]:
        n = _count_for(conn, name)
        if n is not None:
            best = n if best is None else max(best, n)

    if best is None:
        return "none"
    if best >= HEAVY_THRESHOLD:
        return "heavy"
    return "yes" if best >= YES_THRESHOLD else "low"


def stats(conn) -> dict:
    row = conn.execute(
        "SELECT COUNT(*) c, SUM(count) t, MAX(updated_at) u FROM sponsor_records"
    ).fetchone()
    return {"companies": row["c"] or 0, "records": row["t"] or 0, "updated": row["u"]}
